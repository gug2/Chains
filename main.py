import sys
import gui
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.Qt import *
import chains

# == PY QT -- MODULE == #
def createTable(self, matrix):
    table = QTableWidget(self)
    table.setRowCount(len(matrix))
    # get max columns in matrix
    maxColumns = max(len(matrix[i]) for i in range(len(matrix)))
    table.setColumnCount(maxColumns)
    
    for i in range(table.rowCount()):
        for j in range(len(matrix[i])):
            item = QTableWidgetItem(str(matrix[i][j]))
            item.setBackground(QBrush(QColor(255, 255, 255)))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            table.setItem(i, j, item)
    
    table.resizeRowsToContents()
    table.resizeColumnsToContents()
    
    return table
# == PY QT -- END OF MODULE == #


# == OPEN PY XL -- MODULE == #
import openpyxl as pyxl
import random as rand

def clearCellsColor(sheet):
    startCell = getStartCell(sheet)
    
    for row in sheet.iter_rows(startCell.row, sheet.max_row, startCell.column, sheet.max_column):
        for cell in row:
            cell.fill = pyxl.styles.PatternFill(start_color=None)

def getStartCell(sheet):
    for i in range(1, sheet.max_row): #because row starts with 1
        for j in range(0, sheet.max_column):
            if sheet[i][j].value != None:
                return sheet[i][j]
    
    return None

#deprecated
def toMatrix(sheet):
    startCell = getStartCell(sheet)
    
    matrix = []
    for row in sheet.iter_rows(startCell.row, sheet.max_row, startCell.column, sheet.max_column):
        matrixRow = []
        for cell in row:
            if cell.value != None:
                matrixRow.append(cell.value)
        if matrixRow:
            matrix.append(matrixRow)
    
    return matrix

def markChain(sheet, chain, matrix):
    startCell = getStartCell(sheet)
    
    r = rand.randint(64, 255)
    g = rand.randint(64, 255)
    b = rand.randint(64, 255)
    randColor = hex(255 << 24 | r << 16 | g << 8 | b)[2::]

    sheetMaxColumn = sheet.max_column
    for tupl in chain:
        i, j, elem = tupl
        # invers indexes
        i = len(matrix)-1 - i
        i += 1 # because rows starts with 1
        # offset for start cell
        i += startCell.row-1
        j += startCell.column-1
        #print(randColor)
        pattern = pyxl.styles.PatternFill(start_color=randColor, fill_type='solid')
        sheet[i][j].fill = pattern
        result = sheet.cell(row=i, column=sheetMaxColumn+1)
        result.value = sheet[i][j].value
        result.fill = pattern
        
# == OPEN PY XL -- END OF MODULE == #


class Main(QMainWindow, gui.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        workbook = pyxl.open('sample.xlsx')
        sheet = workbook.active
        clearCellsColor(sheet)

        # 1 table
        matrix = [[1,2,3]]
        self.label_1 = QLabel(self)
        self.label_1.setText('Предсказанные значения')
        self.table1 = createTable(self, matrix)

        # 2 table
        chains.matrix = toMatrix(sheet)
        self.label_2 = QLabel(self)
        self.label_2.setText('Введенные значения')
        self.table2 = createTable(self, chains.matrix)

        # 3 table
        chainsTable = []

        # inverse matrix rows for chain detection
        chains.matrix = chains.matrix[::-1]
        
        for i in range(len(chains.matrix)):
            for j in range(len(chains.matrix[i])):
                chainsarray = chains.chainsFor(chains.matrix[i][j], i, j)
                if chainsarray:
                    for i2 in range(len(chainsarray)):
                        for j2 in range(len(chainsarray[i2])):
                            print(chainsarray[i2][j2])

                            # Separate each chain to new sheet for debug
                            #newSheet = workbook.create_sheet('newsheet' + str(i2*len(chainsarray[i2])+j2))
                            #copy
                            #for row in sheet:
                            #    for cell in row:
                            #        newSheet[cell.coordinate].value = cell.value                
                            # mark cells by colors
                            markChain(sheet, chainsarray[i2][j2], chains.matrix)
                            chainsTable.append(chainsarray[i2][j2])
              
        #save workbook
        workbook.save('sample2.xlsx')
        
        # 3 table
        newT = []
        startCell = getStartCell(sheet)
        # максимальное кол-во столбцов исходной таблицы
        maxChainsLen = max(len(chains.matrix[i]) for i in range(len(chains.matrix)))
        # длина текущего листа
        w2 = pyxl.open('sample2.xlsx')
        s2 = w2.active
        maxSheetLen = s2.max_column
        for row in s2.iter_rows(startCell.row, s2.max_row, maxChainsLen+1, maxSheetLen):
            newR = []
            for cell in row:
                if cell.value != None:
                    newR.append(cell.value)
                else:
                    newR.append('')
            newT.append(newR)
        self.label_3 = QLabel(self)
        self.label_3.setText('Последовательности')
        self.table3 = createTable(self, newT)

        self.importAction.triggered.connect(self.openImportMenu)
        
        #grid
        self.grid = QGridLayout(self.centralwidget)
        self.grid.addWidget(self.label_1, 0, 0)
        self.grid.addWidget(self.table1, 1, 0)
        self.grid.addWidget(self.label_2, 2, 0)
        self.grid.addWidget(self.table2, 3, 0)
        self.grid.addWidget(self.label_3, 0, 1)
        self.grid.addWidget(self.table3, 1, 1, 3, 1)

    #def openImportMenu(self):
    #    return QFileDialog.getOpenFileName(self, 'Импорт', '', '*.xlsx;*.xlsm;*.xltx;*.xltm'))[0]
     
app = QApplication(sys.argv)
window = Main()
window.show()
sys.exit(app.exec())
