import openpyxl as pyxl
import random as rand

def getWorkbook(path):
    return pyxl.open(path)
    
def getSheet(workbook):
    return workbook.active
    
def saveWorkbook(workbook, path):
    workbook.save(path)

def clearCellsColor(sheet):
    startCell = getStartCell(sheet)
    
    for row in sheet.iter_rows(startCell.row, sheet.max_row, startCell.column, sheet.max_column):
        for cell in row:
            cell.fill = pyxl.styles.PatternFill(start_color=None)
      
def getStartCell(sheet):
    for i in range(1, sheet.max_row+1): #because row starts with 1
        for j in range(0, sheet.max_column+1):
            if sheet[i][j] and sheet[i][j].value and type(sheet[i][j].value) == int:
                return sheet[i][j]
    
    return None

def toMatrix(sheet):
    startCell = getStartCell(sheet)
    
    matrix = []
    for row in sheet.iter_rows(startCell.row, sheet.max_row, startCell.column, sheet.max_column):
        matrixRow = []
        for cell in row:
            if cell and type(cell.value) == int:
                matrixRow.append(cell.value)
            else:
                matrixRow.append('')
        if matrixRow:
            matrix.append(matrixRow)
    
    return matrix

def markChain(sheet, chain, startColForChains, continueRows, matrix):
    startCell = getStartCell(sheet)
    
    r = rand.randint(64, 255)
    g = rand.randint(64, 255)
    b = rand.randint(64, 255)
    randColor = hex(255 << 24 | r << 16 | g << 8 | b)[2::] # Remove '0x'
    
    for tupl in chain:
        i, j, elem = tupl
        # invers row indexes
        i = len(matrix)-1 - i
        i += 1 # because rows starts with 1
        # offset for start cell
        i += startCell.row-1
        j += startCell.column-1
        # create cell
        pattern = pyxl.styles.PatternFill(start_color=randColor, fill_type='solid')
        
        sheet[i][j].fill = pattern
        
        # offset for continueChains
        result = sheet.cell(row=i+continueRows, column=startColForChains)
        result.value = sheet[i][j].value
        result.fill = pattern
        
def continueChain(sheet, chain, startColForChains, continueRows, matrix):
    deltaRow = abs(chain[-1][0] - chain[-2][0])
    deltaValue = chain[-1][2] - chain[-2][2]
    
    continueColumn = ['' for i in range(continueRows)]
    elementRow = deltaRow
    elementValue = deltaValue
    
    while elementRow <= continueRows:
        resultedRow = elementRow
        # -1 because rows in excel starts with 1 (prepare for list)
        resultedRow -= 1
        # substract offset before size of table and row of last chain's element
        resultedRow -= len(matrix)-1 - chain[-1][0]
        
        continueColumn[resultedRow] = chain[-1][2] + elementValue
        
        # (prepare for excel, adding 1 to row and inverse table)
        result = sheet.cell(row=(continueRows-1-resultedRow)+1, column=startColForChains)
        result.value = chain[-1][2] + elementValue
        
        elementValue += deltaValue
        elementRow += deltaRow
    
    return continueColumn